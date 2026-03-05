from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi.responses import StreamingResponse
from typing import List
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import models, schemas
from database import get_db

router = APIRouter(prefix="/dashboard", tags=["dashboard"])


@router.get("/summary", response_model=schemas.DashboardSummary)
def get_summary(db: Session = Depends(get_db)):
    orders = db.query(models.Order).all()
    total_purchase = sum(1 for o in orders if o.order_type == "Purchase")
    total_sales = sum(1 for o in orders if o.order_type == "Sale")
    pending = sum(1 for o in orders if o.status in ("Pending", "Partial"))
    completed = sum(1 for o in orders if o.status == "Completed")
    total_volume = sum(o.quantity_ordered for o in orders)
    total_value = sum(o.total_price for o in orders)
    return schemas.DashboardSummary(
        total_purchase_orders=total_purchase,
        total_sales_orders=total_sales,
        pending_orders=pending,
        completed_orders=completed,
        total_volume=total_volume,
        total_transaction_value=total_value,
    )


@router.get("/tolerance-log", response_model=List[schemas.ToleranceLogEntry])
def get_tolerance_log(db: Session = Depends(get_db)):
    orders = db.query(models.Order).filter(models.Order.status == "Completed").all()
    result = []
    for o in orders:
        if o.quantity_ordered == 0:
            continue
        deviation = ((o.quantity_sent - o.quantity_ordered) / o.quantity_ordered) * 100
        within = -10 <= deviation <= 10
        result.append(schemas.ToleranceLogEntry(
            order_id=o.order_id,
            broker=o.broker_name,
            item=o.item_name,
            ordered_quantity=o.quantity_ordered,
            actual_quantity=o.quantity_sent,
            deviation_percent=round(deviation, 2),
            within_tolerance=within,
            unit_type=o.unit_type,
        ))
    return result


@router.get("/export")
def export_excel(db: Session = Depends(get_db)):
    orders = db.query(models.Order).order_by(models.Order.created_at.desc()).all()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Orders ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Orders"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    headers = [
        "Order ID", "Broker Name", "Item Name", "Order Type", "Order Date",
        "Qty Ordered", "Unit", "Price/Unit (₹)", "Total Price (₹)",
        "Qty Sent", "Remaining", "Status"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    status_colors = {"Completed": "27AE60", "Partial": "E67E22", "Pending": "E74C3C"}
    for row_idx, o in enumerate(orders, 2):
        row_data = [
            o.order_id, o.broker_name, o.item_name, o.order_type,
            str(o.order_date), o.quantity_ordered, o.unit_type,
            o.price_per_unit, o.total_price, o.quantity_sent,
            o.remaining_quantity, o.status
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        # Colour status cell
        color = status_colors.get(o.status, "FFFFFF")
        ws.cell(row=row_idx, column=12).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        ws.cell(row=row_idx, column=12).font = Font(bold=True, color="FFFFFF")

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    # ── Sheet 2: Tolerance Log ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Tolerance Log")
    tol_headers = ["Order ID", "Broker", "Item", "Ordered Qty", "Actual Qty",
                   "Deviation %", "Within Tolerance", "Unit"]
    for col, h in enumerate(tol_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    tol_orders = db.query(models.Order).filter(models.Order.status == "Completed").all()
    for row_idx, o in enumerate(tol_orders, 2):
        if o.quantity_ordered == 0:
            continue
        deviation = round(((o.quantity_sent - o.quantity_ordered) / o.quantity_ordered) * 100, 2)
        within = -10 <= deviation <= 10
        sign = "+" if deviation >= 0 else ""
        row_data = [
            o.order_id, o.broker_name, o.item_name, o.quantity_ordered,
            o.quantity_sent, f"{sign}{deviation}%", "Yes" if within else "No", o.unit_type
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        tol_color = "27AE60" if within else "E74C3C"
        for col in [6, 7]:
            ws2.cell(row=row_idx, column=col).fill = PatternFill(
                start_color=tol_color, end_color=tol_color, fill_type="solid"
            )
            ws2.cell(row=row_idx, column=col).font = Font(bold=True, color="FFFFFF")

    for col in ws2.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=charbhuja_orders.xlsx"}
    )
